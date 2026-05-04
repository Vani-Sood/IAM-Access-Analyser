"""Compliance report engine — CIS AWS v1.4, NIST 800-53, SOC 2 CC6, PCI-DSS, ISO 27001."""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Callable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.analysis.graph_builder import GraphData

SUPPORTED_FRAMEWORKS = {"cis", "nist", "soc2", "pci", "iso27001"}

_DANGEROUS_IAM_ACTIONS = {
    "iam:*",
    "iam:passrole",
    "iam:createpolicyversion",
    "iam:attachrolepolicy",
    "iam:putrolepolicy",
    "iam:addusertogroup",
    "iam:createaccesskey",
    "iam:updateassumerolepolicy",
    "sts:assumerole",
}

_STATUS_FILL = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "WARNING": "FFEB9C",
    "NOT_APPLICABLE": "D9D9D9",
}


@dataclass(frozen=True)
class ControlResult:
    control_id: str
    framework: str
    title: str
    description: str
    status: str
    details: list[str]
    risk_level: str

    def to_dict(self) -> dict:
        return {
            "control_id": self.control_id,
            "framework": self.framework,
            "title": self.title,
            "description": self.description,
            "status": self.status,
            "details": list(self.details),
            "risk_level": self.risk_level,
        }


@dataclass
class ComplianceReport:
    analysis_id: int
    framework: str
    generated_at: str
    controls: list[ControlResult]

    @property
    def passed(self) -> int:
        return sum(1 for c in self.controls if c.status == "PASS")

    @property
    def failed(self) -> int:
        return sum(1 for c in self.controls if c.status == "FAIL")

    @property
    def warnings(self) -> int:
        return sum(1 for c in self.controls if c.status == "WARNING")

    @property
    def not_applicable(self) -> int:
        return sum(1 for c in self.controls if c.status == "NOT_APPLICABLE")

    @property
    def score(self) -> float:
        applicable = [c for c in self.controls if c.status != "NOT_APPLICABLE"]
        if not applicable:
            return 100.0
        passing = sum(1 for c in applicable if c.status == "PASS")
        return round(passing / len(applicable) * 100, 1)

    def to_dict(self) -> dict:
        return {
            "analysis_id": self.analysis_id,
            "framework": self.framework,
            "generated_at": self.generated_at,
            "controls": [c.to_dict() for c in self.controls],
            "passed": self.passed,
            "failed": self.failed,
            "warnings": self.warnings,
            "not_applicable": self.not_applicable,
            "score": self.score,
        }

    def render_xlsx(self) -> bytes:
        wb = openpyxl.Workbook()

        _build_summary_sheet(wb, self)
        _build_controls_sheet(wb, self)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def _build_summary_sheet(wb: openpyxl.Workbook, report: ComplianceReport) -> None:
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    rows = [
        ("Analysis ID", report.analysis_id),
        ("Framework", report.framework.upper()),
        ("Generated At", report.generated_at),
        ("Score (%)", report.score),
        ("Total Controls", len(report.controls)),
        ("Passed", report.passed),
        ("Failed", report.failed),
        ("Warnings", report.warnings),
        ("Not Applicable", report.not_applicable),
    ]

    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = header_font
        ws.cell(row=i, column=2, value=value)


def _build_controls_sheet(wb: openpyxl.Workbook, report: ComplianceReport) -> None:
    ws = wb.create_sheet("Controls")

    headers = ["Control ID", "Title", "Status", "Risk Level", "Details"]
    widths = [14, 52, 16, 12, 60]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for row_idx, ctrl in enumerate(report.controls, start=2):
        details_str = "; ".join(ctrl.details) if ctrl.details else ""
        values = [ctrl.control_id, ctrl.title, ctrl.status, ctrl.risk_level, details_str]
        fill_color = _STATUS_FILL.get(ctrl.status, "FFFFFF")
        fill = PatternFill(fill_type="solid", fgColor=fill_color)

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)


class ComplianceChecker:
    def __init__(
        self,
        analysis_id: int,
        policy: dict,
        findings: list[dict],
        privesc_paths: list[dict],
        graph: GraphData,
        risk_score: float,
        severity: str,
    ) -> None:
        self._analysis_id = analysis_id
        self._policy = policy
        self._findings = findings
        self._privesc_paths = privesc_paths
        self._graph = graph
        self._risk_score = risk_score
        self._severity = severity
        self._statements = policy.get("Statement", [])

    def check(self, framework: str) -> ComplianceReport:
        if framework not in SUPPORTED_FRAMEWORKS:
            raise ValueError(f"Unsupported framework: {framework!r}. Choose from {SUPPORTED_FRAMEWORKS}")

        dispatch: dict[str, Callable[[], list[ControlResult]]] = {
            "cis": self._check_cis,
            "nist": self._check_nist,
            "soc2": self._check_soc2,
            "pci": self._check_pci,
            "iso27001": self._check_iso27001,
        }
        controls = dispatch[framework]()
        return ComplianceReport(
            analysis_id=self._analysis_id,
            framework=framework,
            generated_at=datetime.now(timezone.utc).isoformat(),
            controls=controls,
        )

    # ── helpers ───────────────────────────────────────────────────────────────

    def _has_full_admin(self) -> tuple[bool, list[str]]:
        """True if any Allow statement grants Action=* on Resource=*."""
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            resources = _normalize_list(stmt.get("Resource", []))
            if ("*" in actions or "iam:*" in actions) and "*" in resources:
                details.append(f"Statement {i}: Action={actions} Resource=*")
        return bool(details), details

    def _has_wildcard_action(self) -> tuple[bool, list[str]]:
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            if "*" in actions or "iam:*" in actions:
                details.append(f"Statement {i}: Action contains wildcard ({actions})")
        return bool(details), details

    def _has_wildcard_resource_on_allow(self) -> tuple[bool, list[str]]:
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            resources = _normalize_list(stmt.get("Resource", []))
            if "*" in resources:
                actions = _normalize_list(stmt.get("Action", []))
                details.append(f"Statement {i}: Resource=* for actions {actions}")
        return bool(details), details

    def _has_direct_user_principal(self) -> tuple[bool, bool, list[str]]:
        """(has_principal_key, has_user_arn, details)."""
        has_principal = False
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            principal = stmt.get("Principal")
            if principal is None:
                continue
            has_principal = True
            arns: list[str] = []
            if isinstance(principal, str):
                arns = [principal]
            elif isinstance(principal, dict):
                aws = principal.get("AWS", [])
                arns = [aws] if isinstance(aws, str) else list(aws)
            for arn in arns:
                if ":user/" in str(arn):
                    details.append(f"Statement {i}: Principal user ARN {arn}")
        return has_principal, bool(details), details

    def _has_dangerous_findings(self) -> tuple[bool, list[str]]:
        critical_high = [
            f for f in self._findings if f.get("severity") in ("CRITICAL", "HIGH")
        ]
        details = [f["message"] for f in critical_high]
        return bool(details), details

    def _has_privesc(self) -> tuple[bool, list[str]]:
        details = [p.get("description", "") for p in self._privesc_paths]
        return bool(details), details

    def _has_dangerous_actions_on_wildcard(self) -> tuple[bool, list[str]]:
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            resources = _normalize_list(stmt.get("Resource", []))
            if "*" not in resources:
                continue
            actions = _normalize_list(stmt.get("Action", []))
            dangerous = [
                a for a in actions
                if a.lower() in _DANGEROUS_IAM_ACTIONS or a == "*"
            ]
            if dangerous:
                details.append(
                    f"Statement {i}: dangerous actions {dangerous} on Resource=*"
                )
        return bool(details), details

    def _allow_statements_without_condition(
        self, action_filter: set[str] | None = None
    ) -> tuple[bool, list[str]]:
        """Allow statements missing a Condition block.
        If action_filter given, only flag stmts whose actions overlap
        filter (or are wildcard)."""
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            if stmt.get("Condition"):
                continue
            actions = _normalize_list(stmt.get("Action", []))
            if action_filter is not None:
                dangerous = [
                    a for a in actions
                    if a == "*" or a.lower() in action_filter
                ]
                if not dangerous:
                    continue
            details.append(
                f"Statement {i}: Allow with no Condition (actions={actions})"
            )
        return bool(details), details

    def _has_cross_account_principal(self) -> tuple[bool, list[str]]:
        """Flag Principal.AWS ARNs that span multiple distinct AWS account IDs."""
        import re
        account_ids: set[str] = set()
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            principal = stmt.get("Principal")
            if not principal:
                continue
            arns: list[str] = []
            if isinstance(principal, str):
                arns = [principal]
            elif isinstance(principal, dict):
                aws = principal.get("AWS", [])
                arns = [aws] if isinstance(aws, str) else list(aws)
            for arn in arns:
                m = re.match(r"arn:aws[^:]*:iam::(\d+):", str(arn))
                if m:
                    account_ids.add(m.group(1))
                    if len(account_ids) > 1:
                        details.append(
                            f"Statement {i}: cross-account ARN {arn}"
                        )
        return len(account_ids) > 1, details

    def _uses_not_action(self) -> tuple[bool, list[str]]:
        details = [
            f"Statement {i}: uses NotAction={stmt['NotAction']}"
            for i, stmt in enumerate(self._statements)
            if "NotAction" in stmt
        ]
        return bool(details), details

    def _uses_not_resource(self) -> tuple[bool, list[str]]:
        details = [
            f"Statement {i}: Allow uses NotResource={stmt['NotResource']}"
            for i, stmt in enumerate(self._statements)
            if "NotResource" in stmt
            and stmt.get("Effect", "Allow") == "Allow"
        ]
        return bool(details), details

    def _has_action_on_wildcard(
        self, actions: set[str]
    ) -> tuple[bool, list[str]]:
        """Any of `actions` (lowercased) Allowed on Resource=*."""
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            resources = _normalize_list(stmt.get("Resource", []))
            if "*" not in resources:
                continue
            stmt_actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in stmt_actions
                if a == "*" or a.lower() in actions
            ]
            if matched:
                details.append(
                    f"Statement {i}: actions {matched} on Resource=*"
                )
        return bool(details), details

    def _has_any_deny_statement(self) -> bool:
        return any(
            stmt.get("Effect") == "Deny" for stmt in self._statements
        )

    def _has_mixed_read_destructive_same_service(
        self,
    ) -> tuple[bool, list[str]]:
        """Same Allow stmt grants both read (Get/List/Describe) and
        destructive (Delete/Put/Create/Update/Attach) verbs on same
        service prefix."""
        _READ = {"get", "list", "describe", "head"}
        _DESTRUCTIVE = {
            "delete", "put", "create", "update",
            "attach", "detach", "remove",
        }
        details: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            by_service: dict[str, set[str]] = {}
            for a in actions:
                if ":" not in a:
                    continue
                svc, verb = a.lower().split(":", 1)
                by_service.setdefault(svc, set()).add(verb)
            for svc, verbs in by_service.items():
                has_read = any(v.startswith(tuple(_READ)) for v in verbs)
                has_dest = any(
                    v.startswith(tuple(_DESTRUCTIVE)) for v in verbs
                )
                if has_read and has_dest:
                    details.append(
                        f"Statement {i}: service '{svc}' has both read"
                        f" and destructive verbs"
                    )
        return bool(details), details

    def _has_mfa_condition(self, stmt: dict) -> bool:
        """True if stmt Condition has aws:MultiFactorAuthPresent = 'true'."""
        cond = stmt.get("Condition", {})
        for op_val in cond.values():
            if not isinstance(op_val, dict):
                continue
            for key, val in op_val.items():
                if key.lower() == "aws:multifactorauthpresent":
                    return str(val).lower() == "true"
        return False

    # ── CIS AWS Benchmark v1.4 ────────────────────────────────────────────────

    def _check_cis(self) -> list[ControlResult]:
        results: list[ControlResult] = []

        # CIS-1.22 — no full admin policy
        full_admin, fa_details = self._has_full_admin()
        results.append(
            ControlResult(
                control_id="CIS-1.22",
                framework="cis",
                title="Ensure IAM policies do not allow full '*:*' administrative privileges",
                description=(
                    "IAM policies must not grant Action=* or iam:* combined with Resource=*."
                ),
                status="FAIL" if full_admin else "PASS",
                details=fa_details,
                risk_level="HIGH",
            )
        )

        # CIS-1.16 — no direct user principal
        has_principal, has_user, user_details = self._has_direct_user_principal()
        if not has_principal:
            cis16_status = "NOT_APPLICABLE"
            cis16_details: list[str] = ["No Principal key present; resource policy context N/A"]
        elif has_user:
            cis16_status = "FAIL"
            cis16_details = user_details
        else:
            cis16_status = "PASS"
            cis16_details = ["No IAM user ARNs in Principal"]
        results.append(
            ControlResult(
                control_id="CIS-1.16",
                framework="cis",
                title="Ensure IAM policies are not attached directly to IAM users",
                description=(
                    "Assign permissions via groups or roles, not directly to user ARNs in Principal."
                ),
                status=cis16_status,
                details=cis16_details,
                risk_level="MEDIUM",
            )
        )

        # CIS-PRIVESC-1 — no privilege escalation
        has_priv, priv_details = self._has_privesc()
        results.append(
            ControlResult(
                control_id="CIS-PRIVESC-1",
                framework="cis",
                title="Ensure no privilege escalation paths exist",
                description=(
                    "IAM policies must not enable privilege escalation to iam:* or sts:AssumeRole on *."
                ),
                status="FAIL" if has_priv else "PASS",
                details=priv_details if priv_details else ["No privilege escalation paths detected"],
                risk_level="HIGH",
            )
        )

        # CIS-WILD-1 — no wildcard resource on dangerous actions
        has_dang, dang_details = self._has_dangerous_actions_on_wildcard()
        results.append(
            ControlResult(
                control_id="CIS-WILD-1",
                framework="cis",
                title="Avoid wildcard (*) resources for sensitive IAM actions",
                description=(
                    "Restrict dangerous IAM actions to specific resource ARNs."
                ),
                status="FAIL" if has_dang else "PASS",
                details=(
                    dang_details if dang_details
                    else ["No dangerous actions with wildcard resources"]
                ),
                risk_level="HIGH",
            )
        )

        # CIS-CONDITION-1 — high-risk Allow without Condition
        _dangerous_lower = {a.lower() for a in _DANGEROUS_IAM_ACTIONS}
        has_cond, cond_details = self._allow_statements_without_condition(
            action_filter=_dangerous_lower
        )
        results.append(
            ControlResult(
                control_id="CIS-CONDITION-1",
                framework="cis",
                title="High-risk Allow without Condition",
                description=(
                    "All Allow statements granting high-risk IAM/STS actions "
                    "must include a Condition block."
                ),
                status="FAIL" if has_cond else "PASS",
                details=(
                    cond_details if cond_details
                    else ["All high-risk Allow statements have Conditions"]
                ),
                risk_level="HIGH",
            )
        )

        # CIS-CROSS-1 — cross-account principal trust
        has_cross, cross_details = self._has_cross_account_principal()
        has_any_principal = any(
            stmt.get("Principal") is not None for stmt in self._statements
        )
        if has_cross:
            cross_status = "WARNING"
            cross_detail_out = cross_details
        elif has_any_principal:
            cross_status = "PASS"
            cross_detail_out = ["No cross-account principals detected"]
        else:
            cross_status = "NOT_APPLICABLE"
            cross_detail_out = ["No Principal key present in any statement"]
        results.append(
            ControlResult(
                control_id="CIS-CROSS-1",
                framework="cis",
                title="Cross-account principal trust",
                description=(
                    "Flag policies that grant access to principals from "
                    "multiple distinct AWS accounts."
                ),
                status=cross_status,
                details=cross_detail_out,
                risk_level="MEDIUM",
            )
        )

        # CIS-NOTACTION-1 — avoid NotAction use
        has_na, na_details = self._uses_not_action()
        results.append(
            ControlResult(
                control_id="CIS-NOTACTION-1",
                framework="cis",
                title="Avoid NotAction use",
                description=(
                    "NotAction grants access to all actions except those listed, "
                    "which can lead to unintended permissions."
                ),
                status="WARNING" if has_na else "PASS",
                details=na_details if na_details else ["No NotAction usage detected"],
                risk_level="MEDIUM",
            )
        )

        # CIS-NOTRESOURCE-1 — avoid NotResource on Allow
        has_nr, nr_details = self._uses_not_resource()
        results.append(
            ControlResult(
                control_id="CIS-NOTRESOURCE-1",
                framework="cis",
                title="Avoid NotResource on Allow",
                description=(
                    "NotResource on Allow statements grants access to all "
                    "resources except those listed, expanding attack surface."
                ),
                status="FAIL" if has_nr else "PASS",
                details=(
                    nr_details if nr_details
                    else ["No NotResource usage on Allow statements"]
                ),
                risk_level="HIGH",
            )
        )

        # CIS-1.17 — ensure support role access exists
        support_actions = {"support:createcase", "support:*"}
        has_support = any(
            a.lower() in support_actions
            for stmt in self._statements
            if stmt.get("Effect", "Allow") == "Allow"
            for a in _normalize_list(stmt.get("Action", []))
        )
        results.append(
            ControlResult(
                control_id="CIS-1.17",
                framework="cis",
                title="Ensure support role access exists",
                description=(
                    "A support role with access to support:CreateCase should "
                    "exist to handle AWS Support incidents."
                ),
                status="PASS" if has_support else "WARNING",
                details=(
                    ["support:CreateCase access detected"]
                    if has_support
                    else [
                        "No support:CreateCase action found; "
                        "ensure a dedicated support role exists"
                    ]
                ),
                risk_level="LOW",
            )
        )

        # CIS-TLS-1 — S3 access without SecureTransport condition
        # Only check S3 Allow stmts that use Resource=* (wildcard)
        s3_allow_stmts = [
            (i, stmt) for i, stmt in enumerate(self._statements)
            if stmt.get("Effect", "Allow") == "Allow"
            and any(
                a.lower().startswith("s3:")
                for a in _normalize_list(stmt.get("Action", []))
            )
            and "*" in _normalize_list(stmt.get("Resource", []))
        ]
        if not s3_allow_stmts:
            tls_status = "NOT_APPLICABLE"
            tls_details: list[str] = ["No S3 Allow statements found"]
        else:
            tls_fail: list[str] = []
            for i, stmt in s3_allow_stmts:
                cond = stmt.get("Condition", {})
                has_tls = any(
                    k.lower() == "aws:securetransport"
                    for op_val in cond.values()
                    if isinstance(op_val, dict)
                    for k in op_val
                )
                if not has_tls:
                    actions = _normalize_list(stmt.get("Action", []))
                    tls_fail.append(
                        f"Statement {i}: S3 actions {actions} "
                        f"without SecureTransport condition"
                    )
            if tls_fail:
                tls_status = "FAIL"
                tls_details = tls_fail
            else:
                tls_status = "PASS"
                tls_details = [
                    "All S3 Allow statements have SecureTransport condition"
                ]
        results.append(
            ControlResult(
                control_id="CIS-TLS-1",
                framework="cis",
                title="S3 access without SecureTransport condition",
                description=(
                    "S3 Allow statements must enforce HTTPS via "
                    "aws:SecureTransport condition."
                ),
                status=tls_status,
                details=tls_details,
                risk_level="MEDIUM",
            )
        )

        # CIS-IP-1 — admin actions without source IP restriction
        # Reuse CIS-CONDITION-1 result; warning (not FAIL) since IP
        # restriction is best-practice, not mandatory
        results.append(
            ControlResult(
                control_id="CIS-IP-1",
                framework="cis",
                title="Admin actions without source IP restriction",
                description=(
                    "Allow statements for high-risk actions should include "
                    "aws:SourceIp condition as a best practice."
                ),
                status="WARNING" if has_cond else "PASS",
                details=(
                    cond_details if cond_details
                    else ["All high-risk Allow statements have Conditions"]
                ),
                risk_level="MEDIUM",
            )
        )

        # CIS-KMS-1 — KMS key deletion/disabling allowed
        kms_danger_actions = {
            "kms:disablekey", "kms:schedulekeydeletion"
        }
        kms_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a.lower() in kms_danger_actions
            ]
            if matched:
                kms_fail.append(
                    f"Statement {i}: KMS key destruction actions {matched}"
                )
        results.append(
            ControlResult(
                control_id="CIS-KMS-1",
                framework="cis",
                title="KMS key deletion/disabling allowed",
                description=(
                    "Policies must not grant kms:DisableKey or "
                    "kms:ScheduleKeyDeletion — these actions permanently "
                    "destroy encryption keys."
                ),
                status="FAIL" if kms_fail else "PASS",
                details=(
                    kms_fail if kms_fail
                    else ["No KMS key destruction actions detected"]
                ),
                risk_level="HIGH",
            )
        )

        # CIS-EC2-1 — overly broad EC2 access
        has_ec2_wild, ec2_details = self._has_action_on_wildcard({"ec2:*"})
        results.append(
            ControlResult(
                control_id="CIS-EC2-1",
                framework="cis",
                title="Overly broad EC2 access",
                description=(
                    "ec2:* on Resource=* grants full control over all EC2 "
                    "resources and is overly permissive."
                ),
                status="FAIL" if has_ec2_wild else "PASS",
                details=(
                    ec2_details if ec2_details
                    else ["No overly broad EC2 access detected"]
                ),
                risk_level="MEDIUM",
            )
        )

        # CIS-PASS-1 — iam:PassRole on wildcard without conditions
        has_pass, pass_details = self._allow_statements_without_condition(
            action_filter={"iam:passrole"}
        )
        results.append(
            ControlResult(
                control_id="CIS-PASS-1",
                framework="cis",
                title="iam:PassRole on wildcard without conditions",
                description=(
                    "iam:PassRole without a Condition allows attackers to "
                    "pass any role to any service, enabling privilege "
                    "escalation."
                ),
                status="FAIL" if has_pass else "PASS",
                details=(
                    pass_details if pass_details
                    else ["No unconditioned iam:PassRole on wildcard detected"]
                ),
                risk_level="HIGH",
            )
        )

        return results

    # ── NIST 800-53 ───────────────────────────────────────────────────────────

    def _check_nist(self) -> list[ControlResult]:
        results: list[ControlResult] = []

        # AC-6 — Least Privilege: no wildcard action
        has_wild, wild_details = self._has_wildcard_action()
        results.append(
            ControlResult(
                control_id="AC-6",
                framework="nist",
                title="Least Privilege",
                description=(
                    "Grant only the minimum permissions necessary. "
                    "Wildcard actions (*) violate least-privilege."
                ),
                status="FAIL" if has_wild else "PASS",
                details=wild_details if wild_details else ["No wildcard actions detected"],
                risk_level="HIGH",
            )
        )

        # AC-6(1) — Authorize access to security functions
        has_dang_wild, dw_details = self._has_dangerous_actions_on_wildcard()
        has_dang_findings, df_details = self._has_dangerous_findings()
        ac61_fail = has_dang_wild or has_dang_findings
        ac61_details = dw_details + df_details
        results.append(
            ControlResult(
                control_id="AC-6(1)",
                framework="nist",
                title="Least Privilege | Authorize Access to Security Functions",
                description=(
                    "Explicitly authorize access to security-sensitive IAM actions "
                    "and restrict them to specific resources."
                ),
                status="FAIL" if ac61_fail else "PASS",
                details=ac61_details if ac61_details else ["No unauthorized security function access"],
                risk_level="HIGH",
            )
        )

        # AC-6(10) — Prohibit privilege escalation
        has_priv, priv_details = self._has_privesc()
        results.append(
            ControlResult(
                control_id="AC-6(10)",
                framework="nist",
                title="Least Privilege | Prohibit Non-Privileged Users from Executing Privileged Functions",
                description="IAM policies must not allow privilege escalation chains.",
                status="FAIL" if has_priv else "PASS",
                details=priv_details if priv_details else ["No privilege escalation paths detected"],
                risk_level="HIGH",
            )
        )

        # AU-2 — Event logging: check for explicit Deny on logging actions
        # N/A for most identity policies; mark based on wildcard deny of logs
        log_actions = {"cloudtrail:*", "logs:*", "cloudwatch:*"}
        deny_log: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect") != "Deny":
                continue
            actions = {a.lower() for a in _normalize_list(stmt.get("Action", []))}
            blocked = actions & log_actions
            if blocked:
                deny_log.append(f"Statement {i} denies logging actions: {blocked}")
        results.append(
            ControlResult(
                control_id="AU-2",
                framework="nist",
                title="Event Logging",
                description="Ensure policies do not block cloud-trail or logging service actions.",
                status="FAIL" if deny_log else "PASS",
                details=deny_log if deny_log else ["No logging actions are explicitly denied"],
                risk_level="MEDIUM",
            )
        )

        # SI-2 — Flaw Remediation: risk-score gate
        if self._risk_score >= 8.0:
            si2_status = "FAIL"
            si2_details = [
                f"Overall risk score {self._risk_score:.1f} ≥ 8"
                f" — immediate remediation required"
            ]
        elif self._risk_score >= 5.0:
            si2_status = "WARNING"
            si2_details = [
                f"Overall risk score {self._risk_score:.1f}"
                f" — review recommended"
            ]
        else:
            si2_status = "PASS"
            si2_details = [
                f"Overall risk score {self._risk_score:.1f}"
                f" within acceptable range"
            ]
        results.append(
            ControlResult(
                control_id="SI-2",
                framework="nist",
                title="Flaw Remediation",
                description=(
                    "Identify, report, and correct information system flaws."
                ),
                status=si2_status,
                details=si2_details,
                risk_level="MEDIUM",
            )
        )

        # AC-2 — Account management: unrestricted user management
        user_mgmt_actions = {
            "iam:createuser", "iam:deleteuser", "iam:updateuser"
        }
        has_user_mgmt, user_mgmt_details = self._has_action_on_wildcard(
            user_mgmt_actions
        )
        results.append(
            ControlResult(
                control_id="AC-2",
                framework="nist",
                title="Account Management",
                description=(
                    "Unrestricted IAM user management actions on Resource=* "
                    "violate account management controls."
                ),
                status="FAIL" if has_user_mgmt else "PASS",
                details=(
                    user_mgmt_details if user_mgmt_details
                    else ["No unrestricted user management actions detected"]
                ),
                risk_level="HIGH",
            )
        )

        # AC-3 — Access enforcement: no explicit deny guards
        has_deny = self._has_any_deny_statement()
        results.append(
            ControlResult(
                control_id="AC-3",
                framework="nist",
                title="Access Enforcement",
                description=(
                    "Policies should include explicit Deny statements to "
                    "enforce least-privilege boundaries."
                ),
                status="PASS" if has_deny else "WARNING",
                details=(
                    ["At least one explicit Deny statement found"]
                    if has_deny
                    else [
                        "No Deny statements found; consider adding explicit "
                        "denies to reinforce access controls"
                    ]
                ),
                risk_level="MEDIUM",
            )
        )

        # AC-5 — Separation of duties
        has_mixed, mixed_details = (
            self._has_mixed_read_destructive_same_service()
        )
        results.append(
            ControlResult(
                control_id="AC-5",
                framework="nist",
                title="Separation of Duties",
                description=(
                    "A single Allow statement should not grant both read and "
                    "destructive verbs on the same service, violating "
                    "separation of duties."
                ),
                status="WARNING" if has_mixed else "PASS",
                details=(
                    mixed_details if mixed_details
                    else ["No mixed read/destructive access detected"]
                ),
                risk_level="MEDIUM",
            )
        )

        # IA-5 — Authenticator management: unrestricted access key ops
        access_key_actions = {
            "iam:createaccesskey",
            "iam:updateaccesskey",
            "iam:deleteaccesskey",
        }
        has_ak, ak_details = self._has_action_on_wildcard(access_key_actions)
        results.append(
            ControlResult(
                control_id="IA-5",
                framework="nist",
                title="Authenticator Management",
                description=(
                    "IAM access key operations on Resource=* allow "
                    "unrestricted credential management, violating "
                    "authenticator controls."
                ),
                status="FAIL" if has_ak else "PASS",
                details=(
                    ak_details if ak_details
                    else ["No unrestricted access key management actions"]
                ),
                risk_level="HIGH",
            )
        )

        # SC-28 — Protection of information at rest
        s3_read_stmts = [
            (i, stmt) for i, stmt in enumerate(self._statements)
            if stmt.get("Effect", "Allow") == "Allow"
            and any(
                a.lower() in {"s3:getobject", "s3:*"}
                for a in _normalize_list(stmt.get("Action", []))
            )
            and "*" in _normalize_list(stmt.get("Resource", []))
        ]
        if not s3_read_stmts:
            sc28_status = "NOT_APPLICABLE"
            sc28_details: list[str] = [
                "No s3:GetObject or s3:* Allow on Resource=* found"
            ]
        else:
            sc28_fail: list[str] = []
            for i, stmt in s3_read_stmts:
                cond = stmt.get("Condition", {})
                has_enc = any(
                    k.lower() == "s3:x-amz-server-side-encryption"
                    for op_val in cond.values()
                    if isinstance(op_val, dict)
                    for k in op_val
                )
                if not has_enc:
                    actions = _normalize_list(stmt.get("Action", []))
                    sc28_fail.append(
                        f"Statement {i}: S3 actions {actions} on * without "
                        f"s3:x-amz-server-side-encryption condition"
                    )
            if sc28_fail:
                sc28_status = "FAIL"
                sc28_details = sc28_fail
            else:
                sc28_status = "PASS"
                sc28_details = [
                    "All S3 read Allows have encryption condition"
                ]
        results.append(
            ControlResult(
                control_id="SC-28",
                framework="nist",
                title="Protection of Information at Rest",
                description=(
                    "S3 read access without server-side encryption condition "
                    "may expose unencrypted data."
                ),
                status=sc28_status,
                details=sc28_details,
                risk_level="MEDIUM",
            )
        )

        # SC-8 — Data in transit protection (NIST framing of TLS check)
        s3_allow_stmts_nist = [
            (i, stmt) for i, stmt in enumerate(self._statements)
            if stmt.get("Effect", "Allow") == "Allow"
            and any(
                a.lower().startswith("s3:")
                for a in _normalize_list(stmt.get("Action", []))
            )
        ]
        if not s3_allow_stmts_nist:
            sc8_status = "PASS"
            sc8_details: list[str] = ["No S3 Allow statements found"]
        else:
            sc8_fail: list[str] = []
            for i, stmt in s3_allow_stmts_nist:
                cond = stmt.get("Condition", {})
                has_tls = any(
                    k.lower() == "aws:securetransport"
                    for op_val in cond.values()
                    if isinstance(op_val, dict)
                    for k in op_val
                )
                if not has_tls:
                    actions = _normalize_list(stmt.get("Action", []))
                    sc8_fail.append(
                        f"Statement {i}: S3 actions {actions} "
                        f"without aws:SecureTransport condition"
                    )
            if sc8_fail:
                sc8_status = "WARNING"
                sc8_details = sc8_fail
            else:
                sc8_status = "PASS"
                sc8_details = [
                    "All S3 Allow statements enforce SecureTransport"
                ]
        results.append(
            ControlResult(
                control_id="SC-8",
                framework="nist",
                title="Data in Transit Protection",
                description=(
                    "S3 Allow statements must enforce HTTPS via "
                    "aws:SecureTransport to protect data in transit."
                ),
                status=sc8_status,
                details=sc8_details,
                risk_level="MEDIUM",
            )
        )

        # IA-2 — Identification and authentication: MFA on high-risk
        _dangerous_lower = {a.lower() for a in _DANGEROUS_IAM_ACTIONS}
        ia2_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            resources = _normalize_list(stmt.get("Resource", []))
            if "*" not in resources:
                continue
            actions = _normalize_list(stmt.get("Action", []))
            dangerous = [
                a for a in actions
                if a == "*" or a.lower() in _dangerous_lower
            ]
            if dangerous and not self._has_mfa_condition(stmt):
                ia2_fail.append(
                    f"Statement {i}: dangerous actions {dangerous} on * "
                    f"without MFA condition"
                )
        results.append(
            ControlResult(
                control_id="IA-2",
                framework="nist",
                title="Identification and Authentication — MFA on High-Risk",
                description=(
                    "Allow statements granting dangerous actions on Resource=* "
                    "must require MFA via aws:MultiFactorAuthPresent condition."
                ),
                status="FAIL" if ia2_fail else "PASS",
                details=(
                    ia2_fail if ia2_fail
                    else [
                        "All dangerous actions on * have MFA condition "
                        "or no such actions exist"
                    ]
                ),
                risk_level="HIGH",
            )
        )

        # CM-5 — Change management: policy modification without MFA
        cm5_danger_actions = {
            "iam:createpolicyversion",
            "iam:putrolepolicy",
            "iam:attachrolepolicy",
            "iam:putuserpolicy",
        }
        cm5_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            resources = _normalize_list(stmt.get("Resource", []))
            if "*" not in resources:
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a.lower() in cm5_danger_actions
            ]
            if matched and not self._has_mfa_condition(stmt):
                cm5_fail.append(
                    f"Statement {i}: policy change actions {matched} on * "
                    f"without MFA condition"
                )
        results.append(
            ControlResult(
                control_id="CM-5",
                framework="nist",
                title="Change Management — Policy Modification without MFA",
                description=(
                    "IAM policy modification actions on Resource=* must "
                    "require MFA to prevent unauthorized configuration changes."
                ),
                status="FAIL" if cm5_fail else "PASS",
                details=(
                    cm5_fail if cm5_fail
                    else [
                        "No unconditioned policy modification actions on * "
                        "detected"
                    ]
                ),
                risk_level="HIGH",
            )
        )

        # AU-9 — Audit log protection
        ct_danger_actions = {
            "cloudtrail:stoplogging",
            "cloudtrail:deletetrail",
            "cloudtrail:updatetrail",
        }
        au9_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a.lower() in ct_danger_actions
            ]
            if matched:
                au9_fail.append(
                    f"Statement {i}: CloudTrail disruption actions {matched}"
                )
        results.append(
            ControlResult(
                control_id="AU-9",
                framework="nist",
                title="Audit Log Protection",
                description=(
                    "Policies must not grant CloudTrail disruption actions "
                    "(StopLogging, DeleteTrail, UpdateTrail) that could "
                    "disable audit logging."
                ),
                status="FAIL" if au9_fail else "PASS",
                details=(
                    au9_fail if au9_fail
                    else ["No CloudTrail disruption actions detected"]
                ),
                risk_level="HIGH",
            )
        )

        # SA-9 — External services: cross-account without conditions
        has_sa9, sa9_details = self._allow_statements_without_condition(
            action_filter={"sts:assumerole"}
        )
        results.append(
            ControlResult(
                control_id="SA-9",
                framework="nist",
                title="External Services — Cross-Account Without Conditions",
                description=(
                    "sts:AssumeRole Allow statements without conditions may "
                    "grant unrestricted cross-account access."
                ),
                status="WARNING" if has_sa9 else "PASS",
                details=(
                    sa9_details if sa9_details
                    else ["All sts:AssumeRole statements have conditions"]
                ),
                risk_level="MEDIUM",
            )
        )

        return results

    # ── SOC 2 CC6 ─────────────────────────────────────────────────────────────

    def _check_soc2(self) -> list[ControlResult]:
        results: list[ControlResult] = []

        # CC6.1 — Logical and physical access controls: risk-score gate
        if self._risk_score >= 8.0:
            cc61_status = "FAIL"
            cc61_details = [f"Risk score {self._risk_score:.1f} indicates critical access control failures"]
        elif self._risk_score >= 6.0:
            cc61_status = "WARNING"
            cc61_details = [f"Risk score {self._risk_score:.1f} indicates elevated access control risk"]
        else:
            cc61_status = "PASS"
            cc61_details = [f"Risk score {self._risk_score:.1f} within acceptable bounds"]
        results.append(
            ControlResult(
                control_id="CC6.1",
                framework="soc2",
                title="Logical and Physical Access Controls",
                description=(
                    "Implement controls that restrict access to information assets to authorized users."
                ),
                status=cc61_status,
                details=cc61_details,
                risk_level="HIGH",
            )
        )

        # CC6.3 — Role-based access: no wildcard resources on Allow
        has_wild_res, wr_details = self._has_wildcard_resource_on_allow()
        results.append(
            ControlResult(
                control_id="CC6.3",
                framework="soc2",
                title="Role-Based Access Control",
                description="Access permissions must be scoped to specific resources, not wildcards.",
                status="FAIL" if has_wild_res else "PASS",
                details=wr_details if wr_details else ["All Allow statements use specific resource ARNs"],
                risk_level="MEDIUM",
            )
        )

        # CC6.6 — Prevent unauthorized access: no privilege escalation
        has_priv, priv_details = self._has_privesc()
        results.append(
            ControlResult(
                control_id="CC6.6",
                framework="soc2",
                title="Prevent Unauthorized Access to Systems",
                description="IAM policies must not enable privilege escalation to broader permissions.",
                status="FAIL" if has_priv else "PASS",
                details=priv_details if priv_details else ["No privilege escalation paths detected"],
                risk_level="HIGH",
            )
        )

        # CC6.7 — Restrict and monitor privileged access
        has_dang_findings, df_details = self._has_dangerous_findings()
        has_dang_wild, dw_details = self._has_dangerous_actions_on_wildcard()
        cc67_fail = has_dang_findings or has_dang_wild
        cc67_details = df_details + dw_details
        results.append(
            ControlResult(
                control_id="CC6.7",
                framework="soc2",
                title="Restrict and Monitor Privileged Access",
                description=(
                    "Privileged IAM actions must be restricted to specific resources "
                    "and monitored for misuse."
                ),
                status="FAIL" if cc67_fail else "PASS",
                details=cc67_details if cc67_details else ["No unrestricted privileged actions detected"],
                risk_level="HIGH",
            )
        )

        # CC6.8 — System component protection: wildcard action check
        has_wild_act, wa_details = self._has_wildcard_action()
        results.append(
            ControlResult(
                control_id="CC6.8",
                framework="soc2",
                title="Protect System Components",
                description=(
                    "Prevent overly broad permissions that expose system "
                    "components."
                ),
                status="WARNING" if has_wild_act else "PASS",
                details=(
                    wa_details if wa_details
                    else ["No wildcard actions detected"]
                ),
                risk_level="MEDIUM",
            )
        )

        # CC6.2 — Authentication: MFA on high-risk actions
        _dangerous_lower = {a.lower() for a in _DANGEROUS_IAM_ACTIONS}
        cc62_target = {"iam:*", "sts:assumerole"}
        cc62_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            resources = _normalize_list(stmt.get("Resource", []))
            if "*" not in resources:
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a == "*"
                or a.lower() in cc62_target
                or a.lower() in _dangerous_lower
            ]
            if matched and not self._has_mfa_condition(stmt):
                cc62_fail.append(
                    f"Statement {i}: high-risk actions {matched} on * "
                    f"without MFA condition"
                )
        results.append(
            ControlResult(
                control_id="CC6.2",
                framework="soc2",
                title="Authentication — MFA on High-Risk Actions",
                description=(
                    "High-risk IAM and STS actions on Resource=* must require "
                    "MFA via aws:MultiFactorAuthPresent condition."
                ),
                status="FAIL" if cc62_fail else "PASS",
                details=(
                    cc62_fail if cc62_fail
                    else [
                        "All high-risk actions on * have MFA condition "
                        "or no such actions exist"
                    ]
                ),
                risk_level="HIGH",
            )
        )

        # CC6.4 — Authorized access revocation
        cc64_actions = {
            "iam:deleteuser", "iam:removeuserfromgroup"
        }
        has_cc64, cc64_details = self._allow_statements_without_condition(
            action_filter=cc64_actions
        )
        results.append(
            ControlResult(
                control_id="CC6.4",
                framework="soc2",
                title="Authorized Access Revocation",
                description=(
                    "IAM user revocation actions without conditions allow "
                    "uncontrolled removal of user access."
                ),
                status="WARNING" if has_cc64 else "PASS",
                details=(
                    cc64_details if cc64_details
                    else [
                        "No unconditioned user revocation actions detected"
                    ]
                ),
                risk_level="MEDIUM",
            )
        )

        # CC6.5 — Logical access: overly broad resource access
        has_ec2_s3, ec2_s3_details = self._has_action_on_wildcard(
            {"ec2:*", "s3:*"}
        )
        results.append(
            ControlResult(
                control_id="CC6.5",
                framework="soc2",
                title="Logical Access — Overly Broad Resource Access",
                description=(
                    "ec2:* or s3:* on Resource=* grants unrestricted access "
                    "to critical infrastructure components."
                ),
                status="WARNING" if has_ec2_s3 else "PASS",
                details=(
                    ec2_s3_details if ec2_s3_details
                    else ["No overly broad EC2 or S3 access detected"]
                ),
                risk_level="MEDIUM",
            )
        )

        # CC7.1 — System monitoring integrity
        ct_actions = {
            "cloudtrail:stoplogging",
            "cloudtrail:deletetrail",
            "cloudtrail:updatetrail",
        }
        cc71_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a.lower() in ct_actions
            ]
            if matched:
                cc71_fail.append(
                    f"Statement {i}: CloudTrail disruption actions {matched}"
                )
        results.append(
            ControlResult(
                control_id="CC7.1",
                framework="soc2",
                title="System Monitoring Integrity",
                description=(
                    "Policies must not grant CloudTrail disruption actions "
                    "that could disable audit logging and monitoring."
                ),
                status="FAIL" if cc71_fail else "PASS",
                details=(
                    cc71_fail if cc71_fail
                    else ["No CloudTrail disruption actions detected"]
                ),
                risk_level="HIGH",
            )
        )

        # CC7.2 — System component evaluation
        kms_destruct = {
            "kms:disablekey", "kms:schedulekeydeletion"
        }
        cc72_fail: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a.lower() in kms_destruct
            ]
            if matched:
                cc72_fail.append(
                    f"Statement {i}: KMS key destruction actions {matched}"
                )
        results.append(
            ControlResult(
                control_id="CC7.2",
                framework="soc2",
                title="System Component Evaluation",
                description=(
                    "KMS key deletion or disabling actions must not be "
                    "permitted as they can destroy encryption infrastructure."
                ),
                status="FAIL" if cc72_fail else "PASS",
                details=(
                    cc72_fail if cc72_fail
                    else ["No KMS key destruction actions detected"]
                ),
                risk_level="HIGH",
            )
        )

        # CC8.1 — Change management
        cc81_actions = {
            "iam:createpolicyversion",
            "iam:putrolepolicy",
            "iam:attachrolepolicy",
        }
        has_cc81, cc81_details = self._has_action_on_wildcard(cc81_actions)
        results.append(
            ControlResult(
                control_id="CC8.1",
                framework="soc2",
                title="Change Management",
                description=(
                    "IAM policy modification actions on Resource=* allow "
                    "unrestricted policy changes without oversight."
                ),
                status="FAIL" if has_cc81 else "PASS",
                details=(
                    cc81_details if cc81_details
                    else ["No unrestricted IAM policy modification detected"]
                ),
                risk_level="HIGH",
            )
        )

        # CC8.2 — Infrastructure changes
        cc82_actions = {
            "ec2:deletevpc",
            "ec2:deletesubnet",
            "ec2:modifyvpcattribute",
        }
        cc82_warn: list[str] = []
        for i, stmt in enumerate(self._statements):
            if stmt.get("Effect", "Allow") != "Allow":
                continue
            actions = _normalize_list(stmt.get("Action", []))
            matched = [
                a for a in actions
                if a.lower() in cc82_actions
            ]
            if matched:
                cc82_warn.append(
                    f"Statement {i}: infrastructure change actions {matched}"
                )
        results.append(
            ControlResult(
                control_id="CC8.2",
                framework="soc2",
                title="Infrastructure Changes",
                description=(
                    "Actions that modify or destroy VPC infrastructure must "
                    "be tightly controlled."
                ),
                status="WARNING" if cc82_warn else "PASS",
                details=(
                    cc82_warn if cc82_warn
                    else ["No uncontrolled infrastructure change actions"]
                ),
                risk_level="MEDIUM",
            )
        )

        # CC9.1 — Risk assessment gate (stricter than CC6.1)
        if self._risk_score >= 9.0:
            cc91_status = "FAIL"
            cc91_details: list[str] = [
                f"Risk score {self._risk_score:.1f} ≥ 9.0 — critical risk "
                f"threshold exceeded"
            ]
        elif self._risk_score >= 7.0:
            cc91_status = "WARNING"
            cc91_details = [
                f"Risk score {self._risk_score:.1f} ≥ 7.0 — elevated risk "
                f"requires review"
            ]
        else:
            cc91_status = "PASS"
            cc91_details = [
                f"Risk score {self._risk_score:.1f} within acceptable bounds"
            ]
        results.append(
            ControlResult(
                control_id="CC9.1",
                framework="soc2",
                title="Risk Assessment Gate",
                description=(
                    "Overall risk score must be below 9.0 (FAIL) and "
                    "below 7.0 (WARNING) to meet SOC 2 risk tolerance."
                ),
                status=cc91_status,
                details=cc91_details,
                risk_level="HIGH",
            )
        )

        # CC9.2 — Risk mitigation: data exfiltration
        has_s3_read, s3_read_details = self._has_action_on_wildcard(
            {"s3:getobject", "s3:*"}
        )
        results.append(
            ControlResult(
                control_id="CC9.2",
                framework="soc2",
                title="Risk Mitigation — Data Exfiltration",
                description=(
                    "s3:GetObject or s3:* on Resource=* creates data "
                    "exfiltration risk that must be mitigated."
                ),
                status="WARNING" if has_s3_read else "PASS",
                details=(
                    s3_read_details if s3_read_details
                    else ["No unrestricted S3 read access detected"]
                ),
                risk_level="MEDIUM",
            )
        )

        return results

    # ── PCI-DSS ───────────────────────────────────────────────────────────────

    def _check_pci(self) -> list[ControlResult]:
        results: list[ControlResult] = []

        # PCI-1.1 — Least Privilege: no wildcard actions
        has_wild, wild_details = self._has_wildcard_action()
        results.append(ControlResult(
            control_id="PCI-1.1",
            framework="pci",
            title="Least Privilege — No Wildcard Actions",
            description="PCI DSS 7.1.2 requires access limited to what is needed. Wildcard actions violate least privilege.",
            status="FAIL" if has_wild else "PASS",
            details=wild_details if has_wild else ["No wildcard actions detected"],
            risk_level="HIGH",
        ))

        # PCI-1.2 — No direct user principal
        has_principal, has_user, user_details = self._has_direct_user_principal()
        if not has_principal:
            pci12_status = "NOT_APPLICABLE"
            pci12_details: list[str] = ["No Principal key in policy"]
        elif has_user:
            pci12_status = "FAIL"
            pci12_details = user_details
        else:
            pci12_status = "PASS"
            pci12_details = ["No direct user ARN principals found"]
        results.append(ControlResult(
            control_id="PCI-1.2",
            framework="pci",
            title="No Direct User Principal",
            description="PCI DSS 7.1 recommends group/role-based access rather than direct user ARN grants.",
            status=pci12_status,
            details=pci12_details,
            risk_level="MEDIUM",
        ))

        # PCI-2.1 — No full admin access
        has_admin, admin_details = self._has_full_admin()
        results.append(ControlResult(
            control_id="PCI-2.1",
            framework="pci",
            title="No Full Admin Access",
            description="PCI DSS 7.1.1 prohibits full administrative access (Action=* Resource=*).",
            status="FAIL" if has_admin else "PASS",
            details=admin_details if has_admin else ["No full admin access detected"],
            risk_level="HIGH",
        ))

        # PCI-3.4 — KMS key protection
        has_kms, kms_details = self._has_action_on_wildcard(
            {"kms:disablekey", "kms:schedulekeydeletion"}
        )
        results.append(ControlResult(
            control_id="PCI-3.4",
            framework="pci",
            title="KMS Key Protection",
            description="PCI DSS 3.5 requires encryption keys to be protected from modification or deletion.",
            status="FAIL" if has_kms else "PASS",
            details=kms_details if has_kms else ["No KMS key deletion/disable on wildcard resource"],
            risk_level="HIGH",
        ))

        # PCI-4.1 — Enforce TLS for S3
        s3_allow_stmts = [
            (i, stmt) for i, stmt in enumerate(self._statements)
            if stmt.get("Effect", "Allow") == "Allow"
            and any(
                a.lower().startswith("s3:")
                for a in _normalize_list(stmt.get("Action", []))
            )
            and "*" in _normalize_list(stmt.get("Resource", []))
        ]
        if not s3_allow_stmts:
            pci41_status = "NOT_APPLICABLE"
            pci41_details: list[str] = ["No S3 Allow statements on wildcard resource"]
        else:
            tls_fail: list[str] = []
            for i, stmt in s3_allow_stmts:
                cond = stmt.get("Condition", {})
                has_tls = any(
                    k.lower() == "aws:securetransport"
                    for op_val in cond.values()
                    if isinstance(op_val, dict)
                    for k in op_val
                )
                if not has_tls:
                    actions = _normalize_list(stmt.get("Action", []))
                    tls_fail.append(
                        f"Statement {i}: S3 actions {actions} without SecureTransport condition"
                    )
            if tls_fail:
                pci41_status = "FAIL"
                pci41_details = tls_fail
            else:
                pci41_status = "PASS"
                pci41_details = ["All S3 Allow statements enforce SecureTransport"]
        results.append(ControlResult(
            control_id="PCI-4.1",
            framework="pci",
            title="Enforce TLS for S3 Data in Transit",
            description="PCI DSS 4.1 requires strong cryptography for data in transit. S3 must enforce HTTPS.",
            status=pci41_status,
            details=pci41_details,
            risk_level="HIGH",
        ))

        # PCI-6.4 — No privilege escalation paths
        has_priv, priv_details = self._has_privesc()
        results.append(ControlResult(
            control_id="PCI-6.4",
            framework="pci",
            title="No Privilege Escalation Paths",
            description="PCI DSS 6.3 requires protection against privilege escalation vulnerabilities.",
            status="FAIL" if has_priv else "PASS",
            details=priv_details if has_priv else ["No privilege escalation paths detected"],
            risk_level="HIGH",
        ))

        # PCI-7.1 — Conditional access for sensitive actions
        has_uncond, uncond_details = self._allow_statements_without_condition(
            {a.lower() for a in _DANGEROUS_IAM_ACTIONS}
        )
        results.append(ControlResult(
            control_id="PCI-7.1",
            framework="pci",
            title="Conditional Access for Sensitive Actions",
            description="PCI DSS 7.2 recommends conditions on high-privilege IAM actions to limit scope.",
            status="WARNING" if has_uncond else "PASS",
            details=uncond_details if has_uncond else ["All high-risk Allow statements have Conditions"],
            risk_level="MEDIUM",
        ))

        # PCI-7.2 — No dangerous actions on wildcard resource
        has_dang, dang_details = self._has_dangerous_actions_on_wildcard()
        results.append(ControlResult(
            control_id="PCI-7.2",
            framework="pci",
            title="No Dangerous Actions on Wildcard Resource",
            description="PCI DSS 7.1.2 prohibits high-risk IAM actions on Resource=*.",
            status="FAIL" if has_dang else "PASS",
            details=dang_details if has_dang else ["No dangerous actions on wildcard resource"],
            risk_level="HIGH",
        ))

        # PCI-8.3 — MFA enforcement check
        high_risk_allow = [
            stmt for stmt in self._statements
            if stmt.get("Effect", "Allow") == "Allow"
            and any(
                a == "*" or a.lower() in _DANGEROUS_IAM_ACTIONS
                for a in _normalize_list(stmt.get("Action", []))
            )
        ]
        mfa_missing = [
            stmt for stmt in high_risk_allow
            if not self._has_mfa_condition(stmt)
        ]
        if not high_risk_allow:
            pci83_status = "NOT_APPLICABLE"
            pci83_details: list[str] = ["No high-risk Allow statements found"]
        elif mfa_missing:
            pci83_status = "WARNING"
            pci83_details = [
                f"Statement {i}: high-risk Allow without MFA condition"
                for i, stmt in enumerate(self._statements)
                if stmt in mfa_missing
            ]
        else:
            pci83_status = "PASS"
            pci83_details = ["All high-risk Allow statements require MFA"]
        results.append(ControlResult(
            control_id="PCI-8.3",
            framework="pci",
            title="MFA Enforcement for High-Risk Actions",
            description="PCI DSS 8.3 requires MFA for administrative and high-privilege access.",
            status=pci83_status,
            details=pci83_details,
            risk_level="MEDIUM",
        ))

        # PCI-10.1 — Audit log protection (CloudTrail)
        has_ct, ct_details = self._has_action_on_wildcard(
            {"cloudtrail:stoplogging", "cloudtrail:deletetrail"}
        )
        results.append(ControlResult(
            control_id="PCI-10.1",
            framework="pci",
            title="Audit Log Protection (CloudTrail)",
            description="PCI DSS 10.5 requires audit logs to be protected from modification or deletion.",
            status="FAIL" if has_ct else "PASS",
            details=ct_details if has_ct else ["No CloudTrail disruption actions on wildcard resource"],
            risk_level="HIGH",
        ))

        # PCI-10.2 — Cross-account access warning
        has_cross, cross_details = self._has_cross_account_principal()
        results.append(ControlResult(
            control_id="PCI-10.2",
            framework="pci",
            title="Cross-Account Access Warning",
            description="PCI DSS 10.3 recommends reviewing cross-account trust relationships.",
            status="WARNING" if has_cross else "PASS",
            details=cross_details if has_cross else ["No cross-account principals detected"],
            risk_level="MEDIUM",
        ))

        return results

    # ── ISO 27001 ─────────────────────────────────────────────────────────────

    def _check_iso27001(self) -> list[ControlResult]:
        results: list[ControlResult] = []

        # ISO-A.9.1 — Access control policy: no wildcard actions
        has_wild, wild_details = self._has_wildcard_action()
        results.append(ControlResult(
            control_id="ISO-A.9.1",
            framework="iso27001",
            title="Access Control Policy — No Wildcard Actions",
            description="ISO 27001 A.9.1 requires an access control policy limiting permissions to what is needed.",
            status="FAIL" if has_wild else "PASS",
            details=wild_details if has_wild else ["No wildcard actions detected"],
            risk_level="HIGH",
        ))

        # ISO-A.9.2 — User access management: no direct user ARN
        has_principal, has_user, user_details = self._has_direct_user_principal()
        if not has_principal:
            iso92_status = "NOT_APPLICABLE"
            iso92_details: list[str] = ["No Principal key in policy"]
        elif has_user:
            iso92_status = "FAIL"
            iso92_details = user_details
        else:
            iso92_status = "PASS"
            iso92_details = ["No direct user ARN principals found"]
        results.append(ControlResult(
            control_id="ISO-A.9.2",
            framework="iso27001",
            title="User Access Management",
            description="ISO 27001 A.9.2 requires formal user registration and de-registration. Avoid direct user ARNs.",
            status=iso92_status,
            details=iso92_details,
            risk_level="MEDIUM",
        ))

        # ISO-A.9.4 — Restriction of access: no full admin
        has_admin, admin_details = self._has_full_admin()
        results.append(ControlResult(
            control_id="ISO-A.9.4",
            framework="iso27001",
            title="Restriction of Access — No Full Admin",
            description="ISO 27001 A.9.4.1 requires restriction of access to information and systems.",
            status="FAIL" if has_admin else "PASS",
            details=admin_details if has_admin else ["No full admin access detected"],
            risk_level="HIGH",
        ))

        # ISO-A.10.1 — Cryptographic key management
        has_kms, kms_details = self._has_action_on_wildcard(
            {"kms:disablekey", "kms:schedulekeydeletion"}
        )
        results.append(ControlResult(
            control_id="ISO-A.10.1",
            framework="iso27001",
            title="Cryptographic Key Management",
            description="ISO 27001 A.10.1.2 requires controls over cryptographic key management including deletion.",
            status="FAIL" if has_kms else "PASS",
            details=kms_details if has_kms else ["No KMS key deletion/disable on wildcard resource"],
            risk_level="HIGH",
        ))

        # ISO-A.12.1 — Broad EC2 access
        has_ec2, ec2_details = self._has_action_on_wildcard(
            {"ec2:*", "ec2:terminateinstances", "ec2:runinstances"}
        )
        results.append(ControlResult(
            control_id="ISO-A.12.1",
            framework="iso27001",
            title="Operational Procedures — Broad EC2 Access",
            description="ISO 27001 A.12.1 requires documented operating procedures. Broad EC2 access lacks control.",
            status="WARNING" if has_ec2 else "PASS",
            details=ec2_details if has_ec2 else ["No overly broad EC2 access detected"],
            risk_level="MEDIUM",
        ))

        # ISO-A.12.4 — Audit log protection (CloudTrail)
        has_ct, ct_details = self._has_action_on_wildcard(
            {"cloudtrail:stoplogging", "cloudtrail:deletetrail"}
        )
        results.append(ControlResult(
            control_id="ISO-A.12.4",
            framework="iso27001",
            title="Audit Log Protection",
            description="ISO 27001 A.12.4 requires event logging and protection of log information.",
            status="FAIL" if has_ct else "PASS",
            details=ct_details if has_ct else ["No CloudTrail disruption actions on wildcard resource"],
            risk_level="HIGH",
        ))

        # ISO-A.12.6 — Technical vulnerability management
        has_dang, dang_details = self._has_dangerous_findings()
        results.append(ControlResult(
            control_id="ISO-A.12.6",
            framework="iso27001",
            title="Technical Vulnerability Management",
            description="ISO 27001 A.12.6 requires timely identification and remediation of technical vulnerabilities.",
            status="FAIL" if has_dang else "PASS",
            details=dang_details if has_dang else ["No CRITICAL/HIGH findings detected"],
            risk_level="HIGH",
        ))

        # ISO-A.13.1 — Network security: TLS enforcement for S3
        s3_allow_stmts = [
            (i, stmt) for i, stmt in enumerate(self._statements)
            if stmt.get("Effect", "Allow") == "Allow"
            and any(
                a.lower().startswith("s3:")
                for a in _normalize_list(stmt.get("Action", []))
            )
            and "*" in _normalize_list(stmt.get("Resource", []))
        ]
        if not s3_allow_stmts:
            iso131_status = "NOT_APPLICABLE"
            iso131_details: list[str] = ["No S3 Allow statements on wildcard resource"]
        else:
            tls_fail2: list[str] = []
            for i, stmt in s3_allow_stmts:
                cond = stmt.get("Condition", {})
                has_tls = any(
                    k.lower() == "aws:securetransport"
                    for op_val in cond.values()
                    if isinstance(op_val, dict)
                    for k in op_val
                )
                if not has_tls:
                    actions = _normalize_list(stmt.get("Action", []))
                    tls_fail2.append(
                        f"Statement {i}: S3 actions {actions} without SecureTransport condition"
                    )
            if tls_fail2:
                iso131_status = "FAIL"
                iso131_details = tls_fail2
            else:
                iso131_status = "PASS"
                iso131_details = ["All S3 Allow statements enforce SecureTransport"]
        results.append(ControlResult(
            control_id="ISO-A.13.1",
            framework="iso27001",
            title="Network Security — TLS Enforcement",
            description="ISO 27001 A.13.1 requires network controls including encryption in transit for S3.",
            status=iso131_status,
            details=iso131_details,
            risk_level="HIGH",
        ))

        # ISO-A.14.2 — Secure development: avoid NotAction
        has_not_action, not_action_details = self._uses_not_action()
        results.append(ControlResult(
            control_id="ISO-A.14.2",
            framework="iso27001",
            title="Secure Development — NotAction Usage",
            description="ISO 27001 A.14.2 requires security in development. NotAction is error-prone and grants unintended access.",
            status="WARNING" if has_not_action else "PASS",
            details=not_action_details if has_not_action else ["No NotAction usage detected"],
            risk_level="MEDIUM",
        ))

        # ISO-A.15.1 — Supplier/third-party access
        has_cross, cross_details = self._has_cross_account_principal()
        results.append(ControlResult(
            control_id="ISO-A.15.1",
            framework="iso27001",
            title="Supplier Relationships — Cross-Account Access",
            description="ISO 27001 A.15.1 requires security controls for third-party access relationships.",
            status="WARNING" if has_cross else "PASS",
            details=cross_details if has_cross else ["No cross-account principals detected"],
            risk_level="MEDIUM",
        ))

        # ISO-A.16.1 — Incident management: dangerous findings
        critical_findings = [
            f for f in self._findings if f.get("severity") == "CRITICAL"
        ]
        high_findings = [
            f for f in self._findings if f.get("severity") == "HIGH"
        ]
        if critical_findings:
            iso161_status = "FAIL"
            iso161_details = [f["message"] for f in critical_findings]
            iso161_risk = "CRITICAL"
        elif high_findings:
            iso161_status = "WARNING"
            iso161_details = [f["message"] for f in high_findings]
            iso161_risk = "HIGH"
        else:
            iso161_status = "PASS"
            iso161_details = ["No CRITICAL or HIGH severity findings"]
            iso161_risk = "LOW"
        results.append(ControlResult(
            control_id="ISO-A.16.1",
            framework="iso27001",
            title="Incident Management — Dangerous Findings",
            description="ISO 27001 A.16.1 requires management of information security incidents.",
            status=iso161_status,
            details=iso161_details,
            risk_level=iso161_risk,
        ))

        # ISO-A.18.1 — Compliance: no privilege escalation
        has_priv, priv_details = self._has_privesc()
        results.append(ControlResult(
            control_id="ISO-A.18.1",
            framework="iso27001",
            title="Compliance — No Privilege Escalation",
            description="ISO 27001 A.18.1 requires compliance with legal requirements. Privilege escalation paths are non-compliant.",
            status="FAIL" if has_priv else "PASS",
            details=priv_details if has_priv else ["No privilege escalation paths detected"],
            risk_level="HIGH",
        ))

        return results


def _normalize_list(value: str | list[str]) -> list[str]:
    if isinstance(value, str):
        return [value]
    return list(value)
