"""Unit tests for compliance engine — Batch 19 (TDD RED phase)."""
from __future__ import annotations

import io

import openpyxl
import pytest

from app.analysis.graph_builder import GraphData


# ── Fixtures ──────────────────────────────────────────────────────────────────

SAFE_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:GetObject"],
            "Resource": "arn:aws:s3:::my-bucket/*",
        }
    ],
}

FULL_ADMIN_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {"Effect": "Allow", "Action": "*", "Resource": "*"},
    ],
}

IAM_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {"Effect": "Allow", "Action": ["iam:*"], "Resource": "*"},
    ],
}

DANGEROUS_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:PassRole", "iam:CreatePolicyVersion"],
            "Resource": "*",
        }
    ],
}

DIRECT_USER_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Principal": {"AWS": "arn:aws:iam::123456789012:user/Alice"},
            "Action": ["s3:GetObject"],
            "Resource": "arn:aws:s3:::my-bucket/*",
        }
    ],
}

SAFE_FINDINGS: list[dict] = []
SAFE_PRIVESC: list[dict] = []

CRITICAL_PRIVESC: list[dict] = [
    {
        "source_node": "stmt-0",
        "target_node": "action-iam:PassRole",
        "path_nodes": ["stmt-0", "action-iam:PassRole"],
        "dangerous_actions": ["iam:PassRole"],
        "severity": "CRITICAL",
        "description": "Privilege escalation via iam:PassRole",
    }
]

HIGH_FINDINGS: list[dict] = [
    {
        "rule_id": "IAM_PASS_ROLE",
        "severity": "HIGH",
        "message": "iam:PassRole detected",
        "statement_idx": 0,
    },
]


def _make_checker(
    policy: dict,
    findings: list[dict] | None = None,
    privesc: list[dict] | None = None,
    graph: GraphData | None = None,
    analysis_id: int = 1,
    risk_score: float = 3.0,
    severity: str = "LOW",
):
    from app.analysis.compliance import ComplianceChecker

    return ComplianceChecker(
        analysis_id=analysis_id,
        policy=policy,
        findings=findings or [],
        privesc_paths=privesc or [],
        graph=graph or GraphData(nodes=[], edges=[]),
        risk_score=risk_score,
        severity=severity,
    )


# ── Construction ──────────────────────────────────────────────────────────────


def test_checker_instantiates():
    assert _make_checker(SAFE_POLICY) is not None


def test_unsupported_framework_raises():
    with pytest.raises(ValueError, match="Unsupported framework"):
        _make_checker(SAFE_POLICY).check("unknown_framework")


# ── ComplianceReport structure ────────────────────────────────────────────────


def test_report_has_required_fields():
    from app.analysis.compliance import ComplianceReport

    report = _make_checker(SAFE_POLICY).check("cis")
    assert isinstance(report, ComplianceReport)
    d = report.to_dict()
    for key in (
        "analysis_id",
        "framework",
        "generated_at",
        "controls",
        "passed",
        "failed",
        "warnings",
        "not_applicable",
        "score",
    ):
        assert key in d, f"missing key: {key}"


def test_report_analysis_id():
    assert _make_checker(SAFE_POLICY, analysis_id=42).check("cis").to_dict()["analysis_id"] == 42


def test_report_framework_field():
    assert _make_checker(SAFE_POLICY).check("nist").to_dict()["framework"] == "nist"
    assert _make_checker(SAFE_POLICY).check("soc2").to_dict()["framework"] == "soc2"


def test_report_controls_is_list():
    assert isinstance(_make_checker(SAFE_POLICY).check("cis").to_dict()["controls"], list)


def test_report_controls_not_empty():
    for fw in ("cis", "nist", "soc2"):
        assert len(_make_checker(SAFE_POLICY).check(fw).to_dict()["controls"]) > 0


def test_control_has_required_fields():
    ctrl = _make_checker(SAFE_POLICY).check("cis").to_dict()["controls"][0]
    for key in ("control_id", "title", "status", "details", "risk_level"):
        assert key in ctrl, f"missing key: {key}"


def test_control_status_valid_values():
    valid = {"PASS", "FAIL", "WARNING", "NOT_APPLICABLE"}
    for fw in ("cis", "nist", "soc2"):
        for ctrl in _make_checker(SAFE_POLICY).check(fw).to_dict()["controls"]:
            assert ctrl["status"] in valid


def test_control_risk_level_valid_values():
    valid = {"HIGH", "MEDIUM", "LOW", "INFO"}
    for fw in ("cis", "nist", "soc2"):
        for ctrl in _make_checker(SAFE_POLICY).check(fw).to_dict()["controls"]:
            assert ctrl["risk_level"] in valid


def test_control_details_is_list():
    for ctrl in _make_checker(SAFE_POLICY).check("cis").to_dict()["controls"]:
        assert isinstance(ctrl["details"], list)


# ── Score ─────────────────────────────────────────────────────────────────────


def test_score_is_float():
    assert isinstance(_make_checker(SAFE_POLICY).check("cis").to_dict()["score"], float)


def test_score_in_range():
    score = _make_checker(SAFE_POLICY).check("cis").to_dict()["score"]
    assert 0.0 <= score <= 100.0


def test_safe_policy_high_score():
    assert _make_checker(SAFE_POLICY).check("cis").to_dict()["score"] >= 75.0


def test_full_admin_low_score():
    score = _make_checker(
        FULL_ADMIN_POLICY, risk_score=9.0, severity="CRITICAL", privesc=CRITICAL_PRIVESC
    ).check("cis").to_dict()["score"]
    assert score < 75.0


# ── CIS-1.22 ─────────────────────────────────────────────────────────────────


def test_cis_full_admin_fails_1_22():
    report = _make_checker(FULL_ADMIN_POLICY, risk_score=9.0, severity="CRITICAL").check("cis")
    ctrl = next(c for c in report.to_dict()["controls"] if "1.22" in c["control_id"])
    assert ctrl["status"] == "FAIL"


def test_cis_iam_wildcard_fails_1_22():
    report = _make_checker(IAM_WILDCARD_POLICY).check("cis")
    ctrl = next(c for c in report.to_dict()["controls"] if "1.22" in c["control_id"])
    assert ctrl["status"] == "FAIL"


def test_cis_safe_passes_1_22():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(c for c in report.to_dict()["controls"] if "1.22" in c["control_id"])
    assert ctrl["status"] == "PASS"


# ── CIS-1.16 ─────────────────────────────────────────────────────────────────


def test_cis_direct_user_principal_fails_1_16():
    report = _make_checker(DIRECT_USER_POLICY).check("cis")
    ctrl = next(c for c in report.to_dict()["controls"] if "1.16" in c["control_id"])
    assert ctrl["status"] == "FAIL"


def test_cis_no_principal_key_is_not_applicable_1_16():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(c for c in report.to_dict()["controls"] if "1.16" in c["control_id"])
    assert ctrl["status"] == "NOT_APPLICABLE"


# ── CIS-PRIVESC ───────────────────────────────────────────────────────────────


def test_cis_privesc_paths_fail():
    report = _make_checker(DANGEROUS_POLICY, privesc=CRITICAL_PRIVESC).check("cis")
    ctrl = next(
        c
        for c in report.to_dict()["controls"]
        if "privesc" in c["control_id"].lower() or "escalat" in c["title"].lower()
    )
    assert ctrl["status"] == "FAIL"


def test_cis_no_privesc_paths_pass():
    report = _make_checker(SAFE_POLICY, privesc=[]).check("cis")
    ctrl = next(
        c
        for c in report.to_dict()["controls"]
        if "privesc" in c["control_id"].lower() or "escalat" in c["title"].lower()
    )
    assert ctrl["status"] == "PASS"


# ── NIST AC-6 ─────────────────────────────────────────────────────────────────


def test_nist_wildcard_action_fails_ac6():
    report = _make_checker(FULL_ADMIN_POLICY).check("nist")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "AC-6")
    assert ctrl["status"] == "FAIL"


def test_nist_safe_passes_ac6():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "AC-6")
    assert ctrl["status"] == "PASS"


def test_nist_dangerous_iam_fails_ac6_1():
    report = _make_checker(IAM_WILDCARD_POLICY, findings=HIGH_FINDINGS).check("nist")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "AC-6(1)")
    assert ctrl["status"] in ("FAIL", "WARNING")


def test_nist_safe_passes_ac6_1():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "AC-6(1)")
    assert ctrl["status"] == "PASS"


def test_nist_privesc_fails_ac6_10():
    report = _make_checker(DANGEROUS_POLICY, privesc=CRITICAL_PRIVESC).check("nist")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "AC-6(10)")
    assert ctrl["status"] == "FAIL"


def test_nist_no_privesc_passes_ac6_10():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "AC-6(10)")
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC6 ─────────────────────────────────────────────────────────────────


def test_soc2_high_risk_fails_cc61():
    report = _make_checker(FULL_ADMIN_POLICY, risk_score=9.0, severity="CRITICAL").check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.1")
    assert ctrl["status"] == "FAIL"


def test_soc2_low_risk_passes_cc61():
    report = _make_checker(SAFE_POLICY, risk_score=2.0, severity="LOW").check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.1")
    assert ctrl["status"] == "PASS"


def test_soc2_medium_risk_warns_cc61():
    report = _make_checker(DANGEROUS_POLICY, risk_score=6.5, severity="HIGH").check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.1")
    assert ctrl["status"] in ("FAIL", "WARNING")


def test_soc2_wildcard_resource_fails_cc63():
    report = _make_checker(FULL_ADMIN_POLICY).check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.3")
    assert ctrl["status"] == "FAIL"


def test_soc2_specific_resource_passes_cc63():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.3")
    assert ctrl["status"] == "PASS"


def test_soc2_privesc_fails_cc66():
    report = _make_checker(DANGEROUS_POLICY, privesc=CRITICAL_PRIVESC).check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.6")
    assert ctrl["status"] == "FAIL"


def test_soc2_no_privesc_passes_cc66():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.6")
    assert ctrl["status"] == "PASS"


def test_soc2_dangerous_findings_fails_cc67():
    report = _make_checker(IAM_WILDCARD_POLICY, findings=HIGH_FINDINGS).check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.7")
    assert ctrl["status"] in ("FAIL", "WARNING")


def test_soc2_safe_passes_cc67():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(c for c in report.to_dict()["controls"] if c["control_id"] == "CC6.7")
    assert ctrl["status"] == "PASS"


# ── Count consistency ─────────────────────────────────────────────────────────


def test_counts_consistent_with_controls():
    d = _make_checker(SAFE_POLICY).check("cis").to_dict()
    controls = d["controls"]
    assert d["passed"] == sum(1 for c in controls if c["status"] == "PASS")
    assert d["failed"] == sum(1 for c in controls if c["status"] == "FAIL")
    assert d["warnings"] == sum(1 for c in controls if c["status"] == "WARNING")
    assert d["not_applicable"] == sum(1 for c in controls if c["status"] == "NOT_APPLICABLE")


def test_full_admin_more_failures_than_safe():
    safe_failed = _make_checker(SAFE_POLICY).check("cis").to_dict()["failed"]
    bad_failed = _make_checker(
        FULL_ADMIN_POLICY, risk_score=9.0, severity="CRITICAL", privesc=CRITICAL_PRIVESC
    ).check("cis").to_dict()["failed"]
    assert bad_failed > safe_failed


# ── XLSX export ───────────────────────────────────────────────────────────────


def test_render_xlsx_returns_bytes():
    assert isinstance(_make_checker(SAFE_POLICY).check("cis").render_xlsx(), bytes)


def test_render_xlsx_not_empty():
    assert len(_make_checker(SAFE_POLICY).check("cis").render_xlsx()) > 0


def test_render_xlsx_valid_zip_magic():
    """XLSX files are ZIP archives; first two bytes are 'PK'."""
    assert _make_checker(SAFE_POLICY).check("cis").render_xlsx()[:2] == b"PK"


def test_render_xlsx_readable_by_openpyxl():
    wb = openpyxl.load_workbook(io.BytesIO(_make_checker(SAFE_POLICY).check("cis").render_xlsx()))
    assert len(wb.sheetnames) >= 2


def test_render_xlsx_summary_sheet_exists():
    wb = openpyxl.load_workbook(io.BytesIO(_make_checker(SAFE_POLICY).check("cis").render_xlsx()))
    assert "Summary" in wb.sheetnames


def test_render_xlsx_controls_sheet_exists():
    wb = openpyxl.load_workbook(io.BytesIO(_make_checker(SAFE_POLICY).check("cis").render_xlsx()))
    assert "Controls" in wb.sheetnames


def test_render_xlsx_controls_sheet_has_header_and_data():
    wb = openpyxl.load_workbook(io.BytesIO(_make_checker(SAFE_POLICY).check("cis").render_xlsx()))
    ws = wb["Controls"]
    assert ws.max_row >= 2  # header + at least one data row


def test_render_xlsx_controls_sheet_has_correct_headers():
    wb = openpyxl.load_workbook(io.BytesIO(_make_checker(SAFE_POLICY).check("nist").render_xlsx()))
    ws = wb["Controls"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Control ID" in headers
    assert "Status" in headers


def test_render_xlsx_all_frameworks():
    for fw in ("cis", "nist", "soc2"):
        report = _make_checker(SAFE_POLICY).check(fw)
        wb = openpyxl.load_workbook(io.BytesIO(report.render_xlsx()))
        assert "Controls" in wb.sheetnames


# ── Framework control-count assertions ────────────────────────────────────────


def test_cis_produces_14_controls():
    report = _make_checker(SAFE_POLICY).check("cis")
    assert len(report.controls) == 14


def test_nist_produces_15_controls():
    report = _make_checker(SAFE_POLICY).check("nist")
    assert len(report.controls) == 15


def test_soc2_produces_14_controls():
    report = _make_checker(SAFE_POLICY).check("soc2")
    assert len(report.controls) == 14


# ── New CIS control policy fixtures ──────────────────────────────────────────

# A policy with a dangerous Allow and NO Condition
_DANGEROUS_NO_CONDITION_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:PassRole", "iam:CreatePolicyVersion"],
            "Resource": "*",
        }
    ],
}

# A policy with a dangerous Allow WITH a Condition (MFA)
_DANGEROUS_WITH_CONDITION_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:PassRole"],
            "Resource": "*",
            "Condition": {
                "Bool": {"aws:MultiFactorAuthPresent": "true"}
            },
        }
    ],
}

# A policy with cross-account principals
_CROSS_ACCOUNT_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Principal": {
                "AWS": [
                    "arn:aws:iam::111111111111:role/RoleA",
                    "arn:aws:iam::222222222222:role/RoleB",
                ]
            },
            "Action": ["sts:AssumeRole"],
            "Resource": "*",
        }
    ],
}

# A policy with single-account principal (no cross-account)
_SINGLE_ACCOUNT_PRINCIPAL_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Principal": {"AWS": "arn:aws:iam::111111111111:role/RoleA"},
            "Action": ["sts:AssumeRole"],
            "Resource": "*",
        }
    ],
}

# A policy that uses NotAction
_NOT_ACTION_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "NotAction": ["s3:DeleteBucket"],
            "Resource": "*",
        }
    ],
}

# A policy that uses NotResource on Allow
_NOT_RESOURCE_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:*"],
            "NotResource": ["arn:aws:s3:::protected-bucket/*"],
        }
    ],
}

# A policy that includes support:CreateCase
_SUPPORT_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["support:CreateCase", "support:DescribeCases"],
            "Resource": "*",
        }
    ],
}

# A policy with s3 Allow on * with NO SecureTransport condition
_S3_NO_TLS_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:GetObject", "s3:PutObject"],
            "Resource": "*",
        }
    ],
}

# A policy with s3 Allow WITH SecureTransport condition
_S3_WITH_TLS_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:GetObject"],
            "Resource": "*",
            "Condition": {
                "Bool": {"aws:SecureTransport": "true"}
            },
        }
    ],
}

# A policy with kms:DisableKey
_KMS_DISABLE_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["kms:DisableKey", "kms:DescribeKey"],
            "Resource": "*",
        }
    ],
}

# A policy with ec2:* on Resource=*
_EC2_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["ec2:*"],
            "Resource": "*",
        }
    ],
}

# A policy with iam:PassRole on * with no condition
_PASSROLE_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:PassRole"],
            "Resource": "*",
        }
    ],
}


# ── CIS-CONDITION-1 ──────────────────────────────────────────────────────────


def test_cis_condition1_fails_for_dangerous_without_condition():
    report = _make_checker(_DANGEROUS_NO_CONDITION_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-CONDITION-1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_cis_condition1_passes_for_dangerous_with_condition():
    report = _make_checker(_DANGEROUS_WITH_CONDITION_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-CONDITION-1"
    )
    assert ctrl["status"] == "PASS"


def test_cis_condition1_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-CONDITION-1"
    )
    assert ctrl["status"] == "PASS"


# ── CIS-CROSS-1 ───────────────────────────────────────────────────────────────


def test_cis_cross1_warns_for_cross_account_principal():
    report = _make_checker(_CROSS_ACCOUNT_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-CROSS-1"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_cis_cross1_passes_for_single_account_principal():
    report = _make_checker(_SINGLE_ACCOUNT_PRINCIPAL_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-CROSS-1"
    )
    assert ctrl["status"] == "PASS"


def test_cis_cross1_not_applicable_for_no_principal():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-CROSS-1"
    )
    assert ctrl["status"] == "NOT_APPLICABLE"


# ── CIS-NOTACTION-1 ───────────────────────────────────────────────────────────


def test_cis_notaction1_warns_for_not_action_usage():
    report = _make_checker(_NOT_ACTION_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-NOTACTION-1"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_cis_notaction1_passes_for_no_not_action():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-NOTACTION-1"
    )
    assert ctrl["status"] == "PASS"


# ── CIS-NOTRESOURCE-1 ─────────────────────────────────────────────────────────


def test_cis_notresource1_fails_for_allow_not_resource():
    report = _make_checker(_NOT_RESOURCE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-NOTRESOURCE-1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_cis_notresource1_passes_for_no_not_resource():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-NOTRESOURCE-1"
    )
    assert ctrl["status"] == "PASS"


# ── CIS-1.17 ──────────────────────────────────────────────────────────────────


def test_cis_117_passes_when_support_action_present():
    report = _make_checker(_SUPPORT_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-1.17"
    )
    assert ctrl["status"] == "PASS"
    assert ctrl["risk_level"] == "LOW"


def test_cis_117_warns_when_no_support_action():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-1.17"
    )
    assert ctrl["status"] == "WARNING"


# ── CIS-TLS-1 ─────────────────────────────────────────────────────────────────


def test_cis_tls1_fails_for_s3_without_secure_transport():
    report = _make_checker(_S3_NO_TLS_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-TLS-1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "MEDIUM"


def test_cis_tls1_passes_for_s3_with_secure_transport():
    report = _make_checker(_S3_WITH_TLS_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-TLS-1"
    )
    assert ctrl["status"] == "PASS"


def test_cis_tls1_not_applicable_for_no_s3():
    # SAFE_POLICY uses s3:GetObject on specific ARN (not wildcard)
    # Verify it is either NOT_APPLICABLE or PASS (no wildcard s3)
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-TLS-1"
    )
    assert ctrl["status"] in ("NOT_APPLICABLE", "PASS")


# ── CIS-IP-1 ──────────────────────────────────────────────────────────────────


def test_cis_ip1_warns_for_dangerous_without_condition():
    report = _make_checker(_DANGEROUS_NO_CONDITION_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-IP-1"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_cis_ip1_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-IP-1"
    )
    assert ctrl["status"] == "PASS"


# ── CIS-KMS-1 ─────────────────────────────────────────────────────────────────


def test_cis_kms1_fails_for_kms_disable_action():
    report = _make_checker(_KMS_DISABLE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-KMS-1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_cis_kms1_passes_for_no_kms_actions():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-KMS-1"
    )
    assert ctrl["status"] == "PASS"


# ── CIS-EC2-1 ─────────────────────────────────────────────────────────────────


def test_cis_ec21_fails_for_ec2_wildcard():
    report = _make_checker(_EC2_WILDCARD_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-EC2-1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "MEDIUM"


def test_cis_ec21_passes_for_no_ec2_wildcard():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-EC2-1"
    )
    assert ctrl["status"] == "PASS"


# ── CIS-PASS-1 ────────────────────────────────────────────────────────────────


def test_cis_pass1_fails_for_passrole_wildcard_no_condition():
    report = _make_checker(_PASSROLE_WILDCARD_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-PASS-1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_cis_pass1_passes_for_passrole_with_condition():
    report = _make_checker(_DANGEROUS_WITH_CONDITION_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-PASS-1"
    )
    assert ctrl["status"] == "PASS"


def test_cis_pass1_passes_for_no_passrole():
    report = _make_checker(SAFE_POLICY).check("cis")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CIS-PASS-1"
    )
    assert ctrl["status"] == "PASS"


# ── New NIST control policy fixtures ─────────────────────────────────────────

# Policy with iam:CreateUser on wildcard (AC-2)
_IAM_CREATE_USER_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:CreateUser", "iam:DeleteUser"],
            "Resource": "*",
        }
    ],
}

# Policy with a Deny statement (AC-3 pass)
_POLICY_WITH_DENY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:GetObject"],
            "Resource": "arn:aws:s3:::my-bucket/*",
        },
        {
            "Effect": "Deny",
            "Action": ["s3:DeleteObject"],
            "Resource": "*",
        },
    ],
}

# Policy that mixes read and destructive in same statement (AC-5)
_MIXED_READ_DESTRUCTIVE_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:GetObject", "s3:DeleteObject", "s3:ListBucket"],
            "Resource": "*",
        }
    ],
}

# Policy with iam:CreateAccessKey on wildcard (IA-5)
_IAM_ACCESS_KEY_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:CreateAccessKey", "iam:DeleteAccessKey"],
            "Resource": "*",
        }
    ],
}

# Policy with s3:GetObject on * without encryption condition (SC-28)
_S3_NO_ENCRYPTION_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["s3:GetObject"],
            "Resource": "*",
        }
    ],
}

# Policy with dangerous action on * with MFA condition (IA-2 pass)
_IAM_PASSROLE_WITH_MFA_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:PassRole"],
            "Resource": "*",
            "Condition": {
                "Bool": {"aws:MultiFactorAuthPresent": "true"}
            },
        }
    ],
}

# Policy with iam:CreatePolicyVersion on * with no MFA (CM-5 fail)
_IAM_POLICY_CHANGE_NO_MFA_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:CreatePolicyVersion", "iam:PutRolePolicy"],
            "Resource": "*",
        }
    ],
}

# Policy allowing cloudtrail:StopLogging (AU-9 fail)
_CLOUDTRAIL_STOP_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": [
                "cloudtrail:StopLogging",
                "cloudtrail:DeleteTrail",
            ],
            "Resource": "*",
        }
    ],
}

# Policy with sts:AssumeRole on * with no Condition (SA-9 warning)
_STS_ASSUMEROLE_NO_CONDITION_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["sts:AssumeRole"],
            "Resource": "*",
        }
    ],
}


# ── NIST AC-2 ─────────────────────────────────────────────────────────────────


def test_nist_ac2_fails_for_iam_user_management_on_wildcard():
    report = _make_checker(_IAM_CREATE_USER_WILDCARD_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AC-2"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_nist_ac2_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AC-2"
    )
    assert ctrl["status"] == "PASS"


# ── NIST AC-3 ─────────────────────────────────────────────────────────────────


def test_nist_ac3_passes_when_deny_exists():
    report = _make_checker(_POLICY_WITH_DENY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AC-3"
    )
    assert ctrl["status"] == "PASS"
    assert ctrl["risk_level"] == "MEDIUM"


def test_nist_ac3_warns_when_no_deny():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AC-3"
    )
    assert ctrl["status"] == "WARNING"


# ── NIST AC-5 ─────────────────────────────────────────────────────────────────


def test_nist_ac5_warns_for_mixed_read_destructive():
    report = _make_checker(_MIXED_READ_DESTRUCTIVE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AC-5"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_nist_ac5_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AC-5"
    )
    assert ctrl["status"] == "PASS"


# ── NIST IA-5 ─────────────────────────────────────────────────────────────────


def test_nist_ia5_fails_for_access_key_on_wildcard():
    report = _make_checker(_IAM_ACCESS_KEY_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "IA-5"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_nist_ia5_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "IA-5"
    )
    assert ctrl["status"] == "PASS"


# ── NIST SC-28 ────────────────────────────────────────────────────────────────


def test_nist_sc28_fails_for_s3_getobject_on_wildcard_no_encryption():
    report = _make_checker(_S3_NO_ENCRYPTION_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "SC-28"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "MEDIUM"


def test_nist_sc28_not_applicable_for_no_s3():
    # Policy with no s3 actions
    no_s3_policy = {
        "Version": "2012-10-17",
        "Statement": [
            {"Effect": "Allow", "Action": ["ec2:DescribeInstances"], "Resource": "*"}
        ],
    }
    report = _make_checker(no_s3_policy).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "SC-28"
    )
    assert ctrl["status"] == "NOT_APPLICABLE"


# ── NIST SC-8 ─────────────────────────────────────────────────────────────────


def test_nist_sc8_warns_for_s3_without_secure_transport():
    report = _make_checker(_S3_NO_TLS_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "SC-8"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_nist_sc8_passes_for_s3_with_secure_transport():
    report = _make_checker(_S3_WITH_TLS_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "SC-8"
    )
    assert ctrl["status"] == "PASS"


# ── NIST IA-2 ─────────────────────────────────────────────────────────────────


def test_nist_ia2_fails_for_dangerous_on_wildcard_without_mfa():
    report = _make_checker(_DANGEROUS_NO_CONDITION_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "IA-2"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_nist_ia2_passes_for_dangerous_with_mfa():
    report = _make_checker(_IAM_PASSROLE_WITH_MFA_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "IA-2"
    )
    assert ctrl["status"] == "PASS"


def test_nist_ia2_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "IA-2"
    )
    assert ctrl["status"] == "PASS"


# ── NIST CM-5 ─────────────────────────────────────────────────────────────────


def test_nist_cm5_fails_for_policy_change_without_mfa():
    report = _make_checker(_IAM_POLICY_CHANGE_NO_MFA_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CM-5"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_nist_cm5_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CM-5"
    )
    assert ctrl["status"] == "PASS"


# ── NIST AU-9 ─────────────────────────────────────────────────────────────────


def test_nist_au9_fails_for_cloudtrail_stop_logging():
    report = _make_checker(_CLOUDTRAIL_STOP_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AU-9"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_nist_au9_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "AU-9"
    )
    assert ctrl["status"] == "PASS"


# ── NIST SA-9 ─────────────────────────────────────────────────────────────────


def test_nist_sa9_warns_for_sts_assumerole_without_condition():
    report = _make_checker(_STS_ASSUMEROLE_NO_CONDITION_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "SA-9"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_nist_sa9_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("nist")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "SA-9"
    )
    assert ctrl["status"] == "PASS"


# ── New SOC 2 control policy fixtures ────────────────────────────────────────

# Policy for CC6.4 — iam:DeleteUser on wildcard with no condition
_IAM_DELETE_USER_NO_CONDITION_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["iam:DeleteUser", "iam:RemoveUserFromGroup"],
            "Resource": "*",
        }
    ],
}

# Policy for CC8.2 — ec2:deleteVpc on wildcard
_EC2_DELETE_VPC_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["ec2:DeleteVpc", "ec2:DeleteSubnet"],
            "Resource": "*",
        }
    ],
}

# Policy for CC6.5 — ec2:* on *
_EC2_S3_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["ec2:*", "s3:*"],
            "Resource": "*",
        }
    ],
}

# Policy for CC8.1 — iam:CreatePolicyVersion on wildcard
_IAM_POLICY_CHANGE_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": [
                "iam:CreatePolicyVersion",
                "iam:PutRolePolicy",
                "iam:AttachRolePolicy",
            ],
            "Resource": "*",
        }
    ],
}


# ── SOC 2 CC6.2 ───────────────────────────────────────────────────────────────


def test_soc2_cc62_fails_for_iam_wildcard_on_wildcard_resource_no_mfa():
    report = _make_checker(IAM_WILDCARD_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.2"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_soc2_cc62_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.2"
    )
    assert ctrl["status"] == "PASS"


def test_soc2_cc62_passes_when_mfa_condition_present():
    report = _make_checker(_IAM_PASSROLE_WITH_MFA_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.2"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC6.4 ───────────────────────────────────────────────────────────────


def test_soc2_cc64_warns_for_iam_deleteuser_no_condition():
    report = _make_checker(_IAM_DELETE_USER_NO_CONDITION_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.4"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_soc2_cc64_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.4"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC6.5 ───────────────────────────────────────────────────────────────


def test_soc2_cc65_warns_for_ec2_s3_wildcard():
    report = _make_checker(_EC2_S3_WILDCARD_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.5"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_soc2_cc65_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC6.5"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC7.1 ───────────────────────────────────────────────────────────────


def test_soc2_cc71_fails_for_cloudtrail_stop():
    report = _make_checker(_CLOUDTRAIL_STOP_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC7.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_soc2_cc71_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC7.1"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC7.2 ───────────────────────────────────────────────────────────────


def test_soc2_cc72_fails_for_kms_disable():
    report = _make_checker(_KMS_DISABLE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC7.2"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_soc2_cc72_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC7.2"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC8.1 ───────────────────────────────────────────────────────────────


def test_soc2_cc81_fails_for_iam_policy_change_on_wildcard():
    report = _make_checker(_IAM_POLICY_CHANGE_WILDCARD_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC8.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_soc2_cc81_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC8.1"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC8.2 ───────────────────────────────────────────────────────────────


def test_soc2_cc82_warns_for_ec2_delete_vpc():
    report = _make_checker(_EC2_DELETE_VPC_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC8.2"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_soc2_cc82_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC8.2"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC9.1 ───────────────────────────────────────────────────────────────


def test_soc2_cc91_fails_for_very_high_risk_score():
    report = _make_checker(FULL_ADMIN_POLICY, risk_score=9.5, severity="CRITICAL").check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC9.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_soc2_cc91_warns_for_elevated_risk_score():
    report = _make_checker(DANGEROUS_POLICY, risk_score=7.5, severity="HIGH").check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC9.1"
    )
    assert ctrl["status"] == "WARNING"


def test_soc2_cc91_passes_for_low_risk_score():
    report = _make_checker(SAFE_POLICY, risk_score=3.0, severity="LOW").check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC9.1"
    )
    assert ctrl["status"] == "PASS"


# ── SOC 2 CC9.2 ───────────────────────────────────────────────────────────────


def test_soc2_cc92_warns_for_s3_getobject_on_wildcard():
    report = _make_checker(_S3_NO_ENCRYPTION_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC9.2"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_soc2_cc92_passes_for_safe_policy():
    report = _make_checker(SAFE_POLICY).check("soc2")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "CC9.2"
    )
    assert ctrl["status"] == "PASS"


# ── Helper method unit tests ──────────────────────────────────────────────────


def test_allow_statements_without_condition_detects_missing_condition():
    checker = _make_checker(_DANGEROUS_NO_CONDITION_POLICY)
    has_issue, details = checker._allow_statements_without_condition(
        action_filter={"iam:passrole", "iam:createpolicyversion"}
    )
    assert has_issue is True
    assert len(details) >= 1


def test_allow_statements_without_condition_passes_when_condition_present():
    checker = _make_checker(_DANGEROUS_WITH_CONDITION_POLICY)
    has_issue, details = checker._allow_statements_without_condition(
        action_filter={"iam:passrole"}
    )
    assert has_issue is False
    assert details == []


def test_has_cross_account_principal_detects_multiple_accounts():
    checker = _make_checker(_CROSS_ACCOUNT_POLICY)
    has_cross, details = checker._has_cross_account_principal()
    assert has_cross is True
    assert len(details) >= 1


def test_has_cross_account_principal_passes_for_single_account():
    checker = _make_checker(_SINGLE_ACCOUNT_PRINCIPAL_POLICY)
    has_cross, details = checker._has_cross_account_principal()
    assert has_cross is False


def test_uses_not_action_detects_not_action():
    checker = _make_checker(_NOT_ACTION_POLICY)
    has_issue, details = checker._uses_not_action()
    assert has_issue is True
    assert len(details) >= 1


def test_uses_not_action_passes_for_normal_action():
    checker = _make_checker(SAFE_POLICY)
    has_issue, details = checker._uses_not_action()
    assert has_issue is False


def test_uses_not_resource_detects_not_resource():
    checker = _make_checker(_NOT_RESOURCE_POLICY)
    has_issue, details = checker._uses_not_resource()
    assert has_issue is True


def test_uses_not_resource_passes_for_normal_resource():
    checker = _make_checker(SAFE_POLICY)
    has_issue, details = checker._uses_not_resource()
    assert has_issue is False


def test_has_action_on_wildcard_detects_action():
    checker = _make_checker(_EC2_WILDCARD_POLICY)
    has_issue, details = checker._has_action_on_wildcard({"ec2:*"})
    assert has_issue is True


def test_has_action_on_wildcard_passes_for_specific_resource():
    checker = _make_checker(SAFE_POLICY)
    has_issue, details = checker._has_action_on_wildcard({"s3:getobject"})
    assert has_issue is False


def test_has_any_deny_statement_detects_deny():
    checker = _make_checker(_POLICY_WITH_DENY)
    assert checker._has_any_deny_statement() is True


def test_has_any_deny_statement_returns_false_for_no_deny():
    checker = _make_checker(SAFE_POLICY)
    assert checker._has_any_deny_statement() is False


def test_has_mixed_read_destructive_same_service_detects_mixed():
    checker = _make_checker(_MIXED_READ_DESTRUCTIVE_POLICY)
    has_issue, details = checker._has_mixed_read_destructive_same_service()
    assert has_issue is True
    assert len(details) >= 1


def test_has_mixed_read_destructive_same_service_passes_for_read_only():
    checker = _make_checker(SAFE_POLICY)
    has_issue, details = checker._has_mixed_read_destructive_same_service()
    assert has_issue is False


def test_has_mfa_condition_detects_mfa():
    checker = _make_checker(SAFE_POLICY)
    stmt_with_mfa = {
        "Effect": "Allow",
        "Action": ["iam:PassRole"],
        "Resource": "*",
        "Condition": {"Bool": {"aws:MultiFactorAuthPresent": "true"}},
    }
    assert checker._has_mfa_condition(stmt_with_mfa) is True


def test_has_mfa_condition_returns_false_without_mfa():
    checker = _make_checker(SAFE_POLICY)
    stmt_no_mfa = {
        "Effect": "Allow",
        "Action": ["iam:PassRole"],
        "Resource": "*",
    }
    assert checker._has_mfa_condition(stmt_no_mfa) is False


# ── PCI-DSS fixtures ──────────────────────────────────────────────────────────

_CLOUDTRAIL_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": [
                "cloudtrail:StopLogging",
                "cloudtrail:DeleteTrail",
            ],
            "Resource": "*",
        }
    ],
}

_KMS_WILDCARD_POLICY = {
    "Version": "2012-10-17",
    "Statement": [
        {
            "Effect": "Allow",
            "Action": ["kms:DisableKey", "kms:ScheduleKeyDeletion"],
            "Resource": "*",
        }
    ],
}

_CRITICAL_FINDINGS_LIST: list[dict] = [
    {
        "rule_id": "FULL_ADMIN",
        "severity": "CRITICAL",
        "message": "Full admin access granted",
        "statement_idx": 0,
    }
]


# ── PCI-DSS ───────────────────────────────────────────────────────────────────


def test_pci_produces_11_controls():
    report = _make_checker(SAFE_POLICY).check("pci")
    assert len(report.controls) == 11


def test_pci_full_admin_fails_pci_21():
    report = _make_checker(FULL_ADMIN_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-2.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_pci_safe_passes_pci_21():
    report = _make_checker(SAFE_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-2.1"
    )
    assert ctrl["status"] == "PASS"


def test_pci_wildcard_action_fails_pci_11():
    report = _make_checker(IAM_WILDCARD_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-1.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_pci_safe_passes_pci_11():
    report = _make_checker(SAFE_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-1.1"
    )
    assert ctrl["status"] == "PASS"


def test_pci_user_principal_fails_pci_12():
    report = _make_checker(DIRECT_USER_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-1.2"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "MEDIUM"


def test_pci_no_user_principal_passes_pci_12():
    report = _make_checker(SAFE_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-1.2"
    )
    assert ctrl["status"] in ("PASS", "NOT_APPLICABLE")


def test_pci_privesc_fails_pci_64():
    report = _make_checker(
        DANGEROUS_POLICY, privesc=CRITICAL_PRIVESC
    ).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-6.4"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_pci_no_privesc_passes_pci_64():
    report = _make_checker(SAFE_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-6.4"
    )
    assert ctrl["status"] == "PASS"


def test_pci_dangerous_wildcard_fails_pci_72():
    report = _make_checker(DANGEROUS_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-7.2"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_pci_safe_passes_pci_72():
    report = _make_checker(SAFE_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-7.2"
    )
    assert ctrl["status"] == "PASS"


def test_pci_cloudtrail_deletion_fails_pci_101():
    report = _make_checker(_CLOUDTRAIL_WILDCARD_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-10.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_pci_cloudtrail_safe_passes_pci_101():
    report = _make_checker(SAFE_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-10.1"
    )
    assert ctrl["status"] == "PASS"


def test_pci_cross_account_warns_pci_102():
    report = _make_checker(_CROSS_ACCOUNT_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-10.2"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_pci_same_account_passes_pci_102():
    report = _make_checker(_SINGLE_ACCOUNT_PRINCIPAL_POLICY).check("pci")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "PCI-10.2"
    )
    assert ctrl["status"] == "PASS"


# ── ISO 27001 ─────────────────────────────────────────────────────────────────


def test_iso27001_produces_12_controls():
    report = _make_checker(SAFE_POLICY).check("iso27001")
    assert len(report.controls) == 12


def test_iso27001_wildcard_action_fails_a91():
    report = _make_checker(IAM_WILDCARD_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.9.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_iso27001_safe_passes_a91():
    report = _make_checker(SAFE_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.9.1"
    )
    assert ctrl["status"] == "PASS"


def test_iso27001_user_principal_fails_a92():
    report = _make_checker(DIRECT_USER_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.9.2"
    )
    assert ctrl["status"] == "FAIL"


def test_iso27001_full_admin_fails_a94():
    report = _make_checker(FULL_ADMIN_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.9.4"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_iso27001_cloudtrail_fails_a124():
    report = _make_checker(_CLOUDTRAIL_WILDCARD_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.12.4"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_iso27001_dangerous_findings_fails_a126():
    report = _make_checker(
        DANGEROUS_POLICY, findings=HIGH_FINDINGS
    ).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.12.6"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_iso27001_safe_passes_a126():
    report = _make_checker(SAFE_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.12.6"
    )
    assert ctrl["status"] == "PASS"


def test_iso27001_not_action_warns_a142():
    report = _make_checker(_NOT_ACTION_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.14.2"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_iso27001_cross_account_warns_a151():
    report = _make_checker(_CROSS_ACCOUNT_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.15.1"
    )
    assert ctrl["status"] == "WARNING"
    assert ctrl["risk_level"] == "MEDIUM"


def test_iso27001_privesc_fails_a181():
    report = _make_checker(
        DANGEROUS_POLICY, privesc=CRITICAL_PRIVESC
    ).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.18.1"
    )
    assert ctrl["status"] == "FAIL"
    assert ctrl["risk_level"] == "HIGH"


def test_iso27001_a161_fails_for_critical_findings():
    report = _make_checker(
        DANGEROUS_POLICY, findings=_CRITICAL_FINDINGS_LIST
    ).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.16.1"
    )
    assert ctrl["status"] == "FAIL"


def test_iso27001_a161_warns_for_high_only_findings():
    report = _make_checker(
        DANGEROUS_POLICY, findings=HIGH_FINDINGS
    ).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.16.1"
    )
    assert ctrl["status"] == "WARNING"


def test_iso27001_a161_passes_for_no_findings():
    report = _make_checker(SAFE_POLICY).check("iso27001")
    ctrl = next(
        c for c in report.to_dict()["controls"]
        if c["control_id"] == "ISO-A.16.1"
    )
    assert ctrl["status"] == "PASS"
